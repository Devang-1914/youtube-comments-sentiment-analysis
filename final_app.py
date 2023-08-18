# ------------------------------------------ IMPORTS --------------------------------------------------#
import openpyxl as openpyxl
from googleapiclient.discovery import build
from datetime import datetime
import streamlit as st
from urllib.parse import urlparse, parse_qs
import pandas as pd
import matplotlib.pyplot as plt
from PIL import Image
from wordcloud import WordCloud, STOPWORDS, ImageColorGenerator
import nltk
import os

nltk.downloader.download('vader_lexicon')
from nltk.sentiment.vader import SentimentIntensityAnalyzer

sentiments = SentimentIntensityAnalyzer()


# ---------------------------------------------- Getting video ID -----------------------------------------#
def get_video_id(videolink):
    parsed_url = urlparse(videolink)
    query_params = parse_qs(parsed_url.query)
    video_id = query_params.get('v', [None])[0]
    return video_id


# --------------------------------------------- Extracting comments --------------------------------------#

def video_comments(youtube, video_id, next_view_token):
    global all_comments, comments_time

    # check for token
    if len(next_view_token.strip()) == 0:
        all_comments = []
        comments_time = []

    if next_view_token == '':
        # get the initial response
        comment_list = youtube.commentThreads().list(part='snippet', maxResults=100, videoId=video_id,
                                                     order='relevance').execute()
        # print(comment_list)
    else:
        # get the next page response
        comment_list = youtube.commentThreads().list(part='snippet', maxResults=100, videoId=video_id,
                                                     order='relevance', pageToken=next_view_token).execute()

    # loop through all top level comments
    for comment in comment_list['items']:
        # add comment to list
        all_comments.append([comment['snippet']['topLevelComment']['snippet']['textDisplay']])
        comments_time.append([comment['snippet']['topLevelComment']['snippet']['publishedAt']])

    if "nextPageToken" in comment_list:
        return video_comments(youtube, video_id, comment_list['nextPageToken'])
    else:
        return []


# ----------------------------------------------- Calling and Storing comments in excel---------------------------------------#

def get_comments_data(v_link):
    global all_comments, comments_time
    os.environ["api_key"] == st.secrets["api_key"]
    all_comments = []
    comments_time = []
    years = []
    months = []

    # build a youtube object using our api key
    yt_object = build('youtube', 'v3', developerKey=api_key)
    video_id = get_video_id(video_link)
    # get all comments and replies
    comments = video_comments(yt_object, video_id, '')

    # Flatten the list of lists to get all values of comment_one in a single list
    all_comment_ones = [comment_one for comment in all_comments for comment_one in comment]
    all_comment_time = [commen_time for comment in comments_time for commen_time in comment]

    st.write(f"The total comments fetched from the video are {len(all_comments)} comments")
    for time in all_comment_time:
        timestamp_str = time
        timestamp = datetime.strptime(timestamp_str, "%Y-%m-%dT%H:%M:%SZ")
        years.append(timestamp.year)
        months.append(timestamp.month)

    # Create a new workbook
    workbook = openpyxl.Workbook()

    # Select the active sheet
    sheet = workbook.active

    # Write the column names to the first row of the sheet
    sheet.cell(row=1, column=1, value="Comment")
    sheet.cell(row=1, column=2, value="Year")
    sheet.cell(row=1, column=3, value="Month")

    # Write the comment_one values to the first column of the sheet
    for idx, comment_one in enumerate(all_comment_ones, start=2):
        sheet.cell(row=idx, column=1, value=comment_one)

    for ydx, comment_year in enumerate(years, start=2):
        sheet.cell(row=ydx, column=2, value=comment_year)

    for mdx, comment_month in enumerate(months, start=2):
        sheet.cell(row=mdx, column=3, value=comment_month)

    # Save the workbook
    excel_name = "comments_video.xlsx"
    workbook.save(excel_name)
    return excel_name


# ------------------------------------------- Sentiment Analysis of Data ------------------------------#

def sentiment_analysis_data(excel_workbook):
    data = pd.read_excel(excel_workbook)
    data['positive'] = [sentiments.polarity_scores(i)["pos"] for i in data['Comment']]
    data['negative'] = [sentiments.polarity_scores(i)['neg'] for i in data['Comment']]
    data['neutral'] = [sentiments.polarity_scores(i)['neu'] for i in data['Comment']]
    x = int(sum(data["positive"]))
    y = int(sum(data["negative"]))
    z = int(sum(data["neutral"]))

    def score(a, b, c):
        if (a > b) and (a > c):
            print("Positive ")
        if (b > a) and (b > c):
            print("Negative")
        if (c > a) and (c > b):
            print("Neutral")

    # st.write(score(x, y, z))
    st.write(f"Positive Score {x}, Negative Score {y}, Neutral Score {z}")

    st.title("Word Cloud Generator")
    text = " ".join(comm for comm in data['Comment'])
    # Generate the word cloud
    stop_words = set(STOPWORDS)
    stop_words.update(["https", "video", "channel", "make", "tutorial", "br", "Ali", "videos"])
    wordcloud = WordCloud(stopwords=stop_words, background_color="white").generate(text)

    # Display the word cloud using Streamlit
    st.image(wordcloud.to_array(), use_column_width=True)
    st.markdown("Word Cloud of Comments")

    st.title("Comments Distribution by Year")

    # Replace 'data' with your actual data source
    comments_per_year = data.groupby('Year')['Comment'].count()

    st.line_chart(comments_per_year, use_container_width=True)



# ------------------------------------------- App Starts ------------------------------------#


# --------------------------------------------- STREAMLIT APP -------------------------------------------#

st.image("youtube-facebook-cover.jpg", caption=None, width=None, use_column_width=None, clamp=False, channels="RGB",
         output_format="auto")
st.title("YouTube Video Comment Analyzer")

# --------------------------------------------- Getting link ---------------------------------------------#
# Get video link from user
video_link = st.text_input("Enter YouTube Video Link:")

if st.button("See Sentiment"):
    comments_data = get_comments_data(video_link)

    sentiment_analysis_data(comments_data)
